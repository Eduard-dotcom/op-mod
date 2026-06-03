#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Comprehensive Archery Muscle Coverage Analysis v2
Cross-references archery anatomical checklist against 11 OpenSim donor models.
"""

import json
import re
import os
from collections import defaultdict

print("=" * 80)
print("АРХЕР: Анализ покрытия мышц моделей OpenSim (v2)")
print("=" * 80)

# ============================================================
# STEP 0: Load all data
# ============================================================
print("\n[Шаг 0] Загрузка данных...")

ARCHERY_FILE = "/home/z/my-project/upload/ПОЛНЫЙ АНАТОМИЧЕСКИЙ ПЕРЕЧЕНЬ ДЛЯ ВЫСТРЕЛА ИЗ ЛУКА (ВЕРСИЯ 2.0).txt"
CATALOG_FILE = "/home/z/my-project/download/muscle_catalog.json"
MUSCLES_FILE = "/home/z/my-project/download/all_muscles_v3.json"
MODELS_FILE  = "/home/z/my-project/download/all_models_v2.json"
OUTPUT_FILE  = "/home/z/my-project/download/Анализ_покрытия_моделей_стрелок.xlsx"

with open(ARCHERY_FILE, 'r', encoding='utf-8') as f:
    archery_text = f.read()
print(f"  Анатомический перечень: {len(archery_text)} символов")

with open(CATALOG_FILE, 'r', encoding='utf-8') as f:
    catalog = json.load(f)
print(f"  Каталог мышц: {len(catalog)} записей")

with open(MUSCLES_FILE, 'r', encoding='utf-8') as f:
    all_muscles = json.load(f)
print(f"  Данные мышц моделей: {list(all_muscles.keys())}")
for mname, mmuscles in all_muscles.items():
    print(f"    {mname}: {len(mmuscles)} мышц")

with open(MODELS_FILE, 'r', encoding='utf-8') as f:
    all_models = json.load(f)

MODEL_NAMES = ["M7", "M7_18", "M8", "M8_Corr", "M8_Norm", "M2", "M4", "M6", "M9", "Raj", "ULB"]

# ============================================================
# STEP 1: Extract muscles from anatomical reference
# ============================================================
print("\n[Шаг 1] Извлечение латинских названий мышц из перечня...")

lines = archery_text.split('\n')
archery_muscles = []
current_phase = "Общее"
current_section = ""

phase_pattern = re.compile(r'ФАЗА\s+[\d\-]+[:\s]', re.IGNORECASE)
section_pattern = re.compile(r'^(\d+\.\d+\.?\s*.+)')

for line in lines:
    line_stripped = line.strip()
    if not line_stripped:
        continue
    if 'ФАЗА' in line_stripped:
        current_phase = line_stripped.strip()
        continue
    sec_match = section_pattern.match(line_stripped)
    if sec_match:
        current_section = sec_match.group(1)
    
    latin_matches = re.findall(r'\(([a-z][a-z\s]+?)\)', line_stripped, re.IGNORECASE)
    for latin in latin_matches:
        latin = latin.strip()
        skip_keywords = [
            'calcaneus', 'talus', 'navicular', 'cuboideum', 'cuneiformes',
            'metatarsalia', 'phalanges', 'tibia', 'fibula', 'femur',
            'sacrum', 'ilium', 'scapula', 'clavicula', 'humerus',
            'metacarpalia', 'subtalar', 'Chopart', 'MTP', 'sacroiliac',
            'sternoclavicular', 'acromioclavicular', 'scapulothoracic',
            'articulatio cubiti', 'fascia', 'thoracolumbar',
        ]
        is_skip = any(kw.lower() in latin.lower() for kw in skip_keywords)
        if not is_skip and len(latin) > 3:
            archery_muscles.append({
                'russian_context': line_stripped,
                'latin_name': latin,
                'phase': current_phase,
                'section': current_section
            })

# Deduplicate
seen_latin = set()
unique_archery_muscles = []
for am in archery_muscles:
    ln = am['latin_name'].lower()
    if ln not in seen_latin:
        seen_latin.add(ln)
        unique_archery_muscles.append(am)

print(f"  Уникальных латинских названий: {len(unique_archery_muscles)}")
for am in unique_archery_muscles:
    print(f"    {am['latin_name']}")

# ============================================================
# STEP 2: Map Latin names to muscle_catalog entries
# ============================================================
print("\n[Шаг 2] Сопоставление латинских названий с каталогом...")

# Build all base_names from catalog for matching
all_base_names = set()
for entry in catalog:
    for bn in entry['base_names']:
        all_base_names.add(bn)

# Map Latin names to catalog entries by matching to base_names
# Strategy: map each Latin name to a list of catalog entry indices
# Use a comprehensive manual mapping based on the actual catalog base_names
LATIN_TO_BASENAMES = {
    # Lower leg
    'tibialis anterior': ['tib_ant', 'tibant'],
    'tibialis posterior': ['tib_post', 'tibpost'],
    'flexor digitorum longus': ['fdl', 'flex_dig'],
    'flexor hallucis longus': ['fhl', 'flex_hal'],
    'peroneus longus': ['per_long', 'perlong'],
    'peroneus brevis': ['per_brev', 'perbrev'],
    
    # Core / Torso
    'diaphragma pelvis': [],  # Not in any model
    'transversus abdominis': [],  # Not in catalog - no TrA entry
    'mm. multifidi': ['MF_m1s', 'MF_m1t_1', 'MF_m1t_2', 'MF_m1t_3',
                       'MF_m2s', 'MF_m2t_1', 'MF_m2t_2', 'MF_m2t_3',
                       'MF_m3s', 'MF_m3t_1', 'MF_m3t_2', 'MF_m3t_3',
                       'MF_m4s', 'MF_m4t_1', 'MF_m4t_2', 'MF_m4t_3',
                       'MF_m5s', 'MF_m5t_1', 'MF_m5t_2', 'MF_m5t_3',
                       'MF_m1_laminar', 'MF_m2_laminar', 'MF_m3_laminar', 
                       'MF_m4_laminar', 'MF_m5_laminar',
                       'deepmult-', 'supmult-', 'multifidus_'],
    'quadratus lumborum': ['QL_ant_', 'QL_mid_', 'QL_post_'],
    'iliocostalis lumborum': ['IL_R5', 'IL_R6', 'IL_R7', 'IL_R8', 'IL_R9', 'IL_R10', 'IL_R11', 'IL_R12',
                               'IL_L1', 'IL_L2', 'IL_L3', 'IL_L4', 'iliocost_cerv_'],
    'longissimus thoracis': ['LTpT_R4', 'LTpT_R5', 'LTpT_R6', 'LTpT_R7', 'LTpT_R8', 
                              'LTpT_R9', 'LTpT_R10', 'LTpT_R11',
                              'LTpT_T1', 'LTpT_T2', 'LTpT_T3', 'LTpT_T4', 'LTpT_T5',
                              'LTpT_T6', 'LTpT_T7', 'LTpT_T8', 'LTpT_T9', 'LTpT_T10', 'LTpT_T11', 'LTpT_T12',
                              'LTpL_'],
    'rectus abdominis': ['rect_abd'],
    'obliquus externus abdominis': ['EO1', 'EO2', 'EO3', 'EO4', 'EO5', 'EO6'],
    'obliquus internus abdominis': ['IO1', 'IO2', 'IO3', 'IO4', 'IO5', 'IO6'],
    'erector spinae': ['LTpT_', 'LTpL_', 'IL_'],
    
    # Shoulder girdle
    'serratus anterior': ['SerrAnt1_1', 'SerrAnt2_1', 'SerrAnt3_1', 'SerrAnt4_1',
                           'SerrAnt5_1', 'SerrAnt6_1', 'SerrAnt7_1', 'SerrAnt8_1', 'SerrAnt9_1'],
    'rhomboideus major et minor': [],  # Not in catalog
    'rhomboidei': [],  # Not in catalog
    'trapezius': ['trap_cl', 'trap_acr_scap', 'trap_acr_T1', 'trap_acr_T2', 'trap_acr_T3',
                   'trap_acr', 'trap_inf_T4', 'trap_inf_T5', 'trap_inf_T6', 'trap_inf_T7',
                   'trap_inf_T8', 'trap_inf_T9', 'trap_inf_T10', 'trap_inf_T11', 'trap_inf_T12'],
    'levator scapulae': ['levator_scap'],
    'deltoideus': ['DELT1', 'DELT2', 'DELT3'],
    'supraspinatus': ['SUPSP'],
    'infraspinatus': ['INFSP'],
    'subscapularis': ['SUBSC'],
    'teres minor': ['TMIN'],
    'teres major': ['TMAJ'],
    
    # Arm
    'biceps brachii caput longum': ['BIClong'],
    'biceps brachii': ['BIClong', 'BICshort'],
    'triceps brachii': ['TRIlong', 'TRIlat', 'TRImed'],
    'anconeus': ['ANC'],
    'coracobrachialis': ['CORB'],
    'brachialis': ['BRA'],
    'brachioradialis': ['BRD'],
    'latissimus dorsi': ['LD_Il', 'LD_L1', 'LD_L2', 'LD_L3', 'LD_L4',
                          'LD_T5', 'LD_T6', 'LD_T7', 'LD_T8', 'LD_T9', 'LD_T10', 'LD_T11', 'LD_T12'],
    
    # Forearm / Hand
    'flexor digitorum profundus': ['FDPI', 'FDPM', 'FDPR', 'FDPL'],
    'flexor digitorum superficialis': ['FDSI', 'FDSM', 'FDSR', 'FDSL'],
    'lumbricales': ['LUMI', 'LUMM', 'LUMR', 'LUML'],
    'interossei': ['1stDI_MC1', '1stDI_MC2', '2ndDI', '3rdDI', '4thDI', '1stPI', '2ndPI', '3rdPI'],
    'flexor pollicis longus': ['FPL'],
    'flexor pollicis longus et brevis': ['FPL', 'FPB'],
    'opponens pollicis': ['OPP'],
    'flexores digitorum': ['FDPI', 'FDPM', 'FDPR', 'FDPL', 'FDSI', 'FDSM', 'FDSR', 'FDSL'],
    'extensores digitorum': ['EDCI', 'EDCM', 'EDCR', 'EDCL', 'EDM', 'EIP', 'edl', 'ext_dig'],
    
    # Hip/pelvis
    'psoas': ['psoas', 'Ps_L1_VB', 'Ps_L1_TP', 'Ps_L2_VB', 'Ps_L2_TP', 
               'Ps_L3_VB', 'Ps_L3_TP', 'Ps_L4_VB', 'Ps_L4_TP', 'Ps_L5_VB', 'Ps_L5_TP',
               'Ps_L1_L2_IVD', 'Ps_L2_L3_IVD', 'Ps_L3_L4_IVD', 'Ps_L4_L5_IVD'],
    'iliacus': ['iliacus'],
}

def match_latin_to_catalog(latin_name):
    """Match a Latin muscle name to catalog entries using the mapping table."""
    ln = latin_name.lower().strip()
    
    # 1. Try exact match in LATIN_TO_BASENAMES
    if ln in LATIN_TO_BASENAMES:
        prefixes = LATIN_TO_BASENAMES[ln]
        if not prefixes:
            return [], "NOT_IN_ANY_MODEL"
        
        matched_entries = []
        matched_indices = set()
        for prefix in prefixes:
            for cat_idx, entry in enumerate(catalog):
                if cat_idx in matched_indices:
                    continue
                for bn in entry['base_names']:
                    if bn == prefix or bn.startswith(prefix):
                        matched_entries.append(entry)
                        matched_indices.add(cat_idx)
                        break
        if matched_entries:
            return matched_entries, "MANUAL_MAP"
        return [], f"MANUAL_MAP: prefixes not found in catalog"
    
    # 2. Try fuzzy matching - search by partial name in base_names
    ln_words = ln.split()
    matched_entries = []
    matched_indices = set()
    
    # Try to find base_names that contain key parts of the latin name
    for cat_idx, entry in enumerate(catalog):
        if cat_idx in matched_indices:
            continue
        for bn in entry['base_names']:
            bn_lower = bn.lower()
            # Check if Latin name appears as underscore-separated in base_name
            if ln.replace(' ', '_') in bn_lower:
                matched_entries.append(entry)
                matched_indices.add(cat_idx)
                break
    
    if matched_entries:
        return matched_entries, "FUZZY_EXACT"
    
    # 3. Try Russian name matching from context
    return [], "NOT_FOUND"

# Perform matching
latin_to_catalog = {}
for am in unique_archery_muscles:
    ln = am['latin_name']
    entries, method = match_latin_to_catalog(ln)
    latin_to_catalog[ln] = {
        'entries': entries,
        'method': method,
        'archery_info': am
    }

found_count = sum(1 for v in latin_to_catalog.values() if v['entries'])
not_found_count = sum(1 for v in latin_to_catalog.values() if not v['entries'])

print(f"  Найдено в каталоге: {found_count}")
print(f"  Не найдено: {not_found_count}")

for ln, info in sorted(latin_to_catalog.items()):
    entries_str = ", ".join(e['muscle_ru'] for e in info['entries'][:3]) if info['entries'] else "—"
    extra = f" (+{len(info['entries'])-3} др.)" if len(info['entries']) > 3 else ""
    print(f"    {ln}: [{info['method']}] → {entries_str}{extra}")

# ============================================================
# STEP 3: Check each model's coverage
# ============================================================
print("\n[Шаг 3] Проверка покрытия мышц в каждой модели...")

# Handle suffixes: _r, _l, _R, _L (some models use uppercase)
def strip_suffix(name):
    """Strip bilateral suffix (_r, _l, _R, _L) from muscle name."""
    for suffix in ('_r', '_l', '_R', '_L'):
        if name.endswith(suffix):
            return name[:-2]
    return name

# Build base_to_count for each model once
print("  Вычисление покрытия...")

model_base_to_count = {}
for model_name in MODEL_NAMES:
    base_to_count = defaultdict(int)
    if model_name in all_muscles:
        for m in all_muscles[model_name]:
            base = strip_suffix(m['name'])
            base_to_count[base] += 1
    model_base_to_count[model_name] = base_to_count

coverage = {}
for cat_idx, entry in enumerate(catalog):
    for model_name in MODEL_NAMES:
        count = 0
        for bn in entry['base_names']:
            if bn in model_base_to_count[model_name]:
                count = max(count, model_base_to_count[model_name][bn])
        coverage[(cat_idx, model_name)] = count

# Print summary
print("\n  Сводка покрытия:")
for model_name in MODEL_NAMES:
    present = sum(1 for (ci, mn), cnt in coverage.items() if mn == model_name and cnt > 0)
    total = len(catalog)
    print(f"    {model_name}: {present}/{total} мышц каталога ({100*present/total:.1f}%)")

# ============================================================
# STEP 4: Build coverage analysis by body region
# ============================================================
print("\n[Шаг 4] Анализ покрытия по отделам тела...")

# Normalize dept names to 4 main regions
def normalize_dept(dept):
    d = dept.strip()
    if d == 'Шея':
        return 'Шея'
    elif d == 'Туловище':
        return 'Туловище'
    elif d in ('Плечевой пояс', 'Плечо', 'Предплечье', 'Кисть'):
        return 'Верхние конечности'
    elif d in ('Таз/Бедро', 'Бедро', 'Голень/Стопа'):
        return 'Нижние конечности'
    else:
        return d

regions = ['Шея', 'Туловище', 'Верхние конечности', 'Нижние конечности']
region_coverage = {}

for region in regions:
    region_entries = []
    for cat_idx, entry in enumerate(catalog):
        if normalize_dept(entry['dept']) == region:
            region_entries.append((cat_idx, entry))
    
    total_muscles = len(region_entries)
    region_coverage[region] = {
        'total': total_muscles,
        'entries': region_entries,
        'by_model': {}
    }
    
    for model_name in MODEL_NAMES:
        present = 0
        missing = []
        for cat_idx, entry in region_entries:
            if coverage.get((cat_idx, model_name), 0) > 0:
                present += 1
            else:
                missing.append(f"{entry['muscle_ru']} ({entry['head_ru']})")
        pct = 100 * present / total_muscles if total_muscles > 0 else 0
        region_coverage[region]['by_model'][model_name] = {
            'present': present,
            'total': total_muscles,
            'pct': pct,
            'missing': missing
        }

for region in regions:
    print(f"\n  {region} ({region_coverage[region]['total']} мышц в каталоге):")
    for model_name in MODEL_NAMES:
        rc = region_coverage[region]['by_model'][model_name]
        if rc['pct'] > 0:
            print(f"    {model_name}: {rc['present']}/{rc['total']} ({rc['pct']:.1f}%)")

# ============================================================
# STEP 5: Identify critical gaps
# ============================================================
print("\n[Шаг 5] Критические дефициты...")

no_model_muscles = []
single_model_muscles = []

for ln, info in latin_to_catalog.items():
    if not info['entries']:
        no_model_muscles.append((ln, info, "НЕ В КАТАЛОГЕ"))
        continue
    
    models_having = set()
    for entry in info['entries']:
        try:
            cat_idx = catalog.index(entry)
        except ValueError:
            continue
        for model_name in MODEL_NAMES:
            if coverage.get((cat_idx, model_name), 0) > 0:
                models_having.add(model_name)
    
    if len(models_having) == 0:
        no_model_muscles.append((ln, info, "В КАТАЛОГЕ, НО НЕТ В МОДЕЛЯХ"))
    elif len(models_having) == 1:
        single_model_muscles.append((ln, info, list(models_having)[0]))

print(f"  Мышцы отсутствующие во ВСЕХ моделях: {len(no_model_muscles)}")
for ln, info, reason in no_model_muscles:
    print(f"    - {ln}: {reason}")

print(f"  Мышцы только в ОДНОЙ модели: {len(single_model_muscles)}")
for ln, info, model in single_model_muscles:
    print(f"    - {ln} (только {model})")

# ============================================================
# STEP 6: Generate XLSX output
# ============================================================
print("\n[Шаг 6] Генерация XLSX файла...")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

wb = Workbook()

# Styles
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
small_font = Font(size=9)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(bold=True, size=11, color='FFFFFF')
green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
light_green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
wrap_align = Alignment(wrap_text=True, vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    return cell

def apply_border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = thin_border

def coverage_fill(pct):
    if pct >= 80: return green_fill
    elif pct >= 50: return yellow_fill
    elif pct > 0: return orange_fill
    else: return red_fill

# ---- Sheet 1: Сводка по отделам ----
print("  Лист 1: Сводка по отделам...")
ws1 = wb.active
ws1.title = "Сводка по отделам"

set_cell(ws1, 1, 1, "СВОДКА ПОКРЫТИЯ МОДЕЛЕЙ ПО ОТДЕЛАМ ТЕЛА", title_font)
ws1.merge_cells('A1:L1')

# Headers
row = 3
for col_idx, h in enumerate(['Отдел тела'] + MODEL_NAMES, 1):
    set_cell(ws1, row, col_idx, h, header_font_white, header_fill, center_align, thin_border)

# Data
for row_idx, region in enumerate(regions, 4):
    set_cell(ws1, row_idx, 1, region, header_font, None, wrap_align, thin_border)
    for col_idx, model_name in enumerate(MODEL_NAMES, 2):
        rc = region_coverage[region]['by_model'][model_name]
        val = f"{rc['present']}/{rc['total']} ({rc['pct']:.1f}%)"
        set_cell(ws1, row_idx, col_idx, val, small_font, coverage_fill(rc['pct']), center_align, thin_border)

# Total row
total_row = 4 + len(regions)
set_cell(ws1, total_row, 1, "ИТОГО", Font(bold=True, size=12), None, wrap_align, thin_border)
for col_idx, model_name in enumerate(MODEL_NAMES, 2):
    total_present = sum(region_coverage[r]['by_model'][model_name]['present'] for r in regions)
    total_expected = sum(region_coverage[r]['total'] for r in regions)
    pct = 100 * total_present / total_expected if total_expected > 0 else 0
    val = f"{total_present}/{total_expected} ({pct:.1f}%)"
    set_cell(ws1, total_row, col_idx, val, Font(bold=True, size=10), coverage_fill(pct), center_align, thin_border)

# Ranking section
rank_start = total_row + 3
set_cell(ws1, rank_start, 1, "РАНЖИРОВАНИЕ МОДЕЛЕЙ ПО ОТДЕЛАМ", title_font)
ws1.merge_cells(start_row=rank_start, start_column=1, end_row=rank_start, end_column=4)

for region_idx, region in enumerate(regions):
    r = rank_start + 2 + region_idx * (2 + len(MODEL_NAMES))
    set_cell(ws1, r, 1, region, Font(bold=True, size=12, color='4472C4'))
    
    model_scores = [(mn, region_coverage[region]['by_model'][mn]['pct'],
                     region_coverage[region]['by_model'][mn]['present'],
                     region_coverage[region]['by_model'][mn]['total'])
                    for mn in MODEL_NAMES]
    model_scores.sort(key=lambda x: x[1], reverse=True)
    
    for ci, h in enumerate(['Место', 'Модель', 'Покрытие', 'мышц'], 1):
        set_cell(ws1, r+1, ci, h, header_font, None, center_align, thin_border)
    
    for rank, (mn, pct, pres, tot) in enumerate(model_scores, 1):
        rr = r + 1 + rank
        fill = green_fill if rank == 1 else (yellow_fill if rank == 2 else None)
        set_cell(ws1, rr, 1, rank, None, fill, center_align, thin_border)
        set_cell(ws1, rr, 2, mn, None, fill, center_align, thin_border)
        set_cell(ws1, rr, 3, f"{pct:.1f}%", None, fill, center_align, thin_border)
        set_cell(ws1, rr, 4, f"{pres}/{tot}", None, fill, center_align, thin_border)

ws1.column_dimensions['A'].width = 25
for i in range(2, len(MODEL_NAMES) + 2):
    ws1.column_dimensions[get_column_letter(i)].width = 14

# ---- Sheet 2: Матрица мышц ----
print("  Лист 2: Матрица мышц...")
ws2 = wb.create_sheet("Матрица мышц")

headers2 = ['№', 'Отдел', 'Подгруппа', 'Мышца', 'Головка'] + MODEL_NAMES + ['В перечне стрелка']
for col_idx, h in enumerate(headers2, 1):
    set_cell(ws2, 1, col_idx, h, header_font_white, header_fill, center_align, thin_border)

# Build archery catalog indices
archery_catalog_indices = set()
for ln, info in latin_to_catalog.items():
    for entry in info['entries']:
        try:
            idx = catalog.index(entry)
            archery_catalog_indices.add(idx)
        except ValueError:
            pass

for cat_idx, entry in enumerate(catalog):
    row = cat_idx + 2
    set_cell(ws2, row, 1, cat_idx + 1, None, None, center_align, thin_border)
    set_cell(ws2, row, 2, entry['dept'], None, None, wrap_align, thin_border)
    set_cell(ws2, row, 3, entry['subgroup'], None, None, wrap_align, thin_border)
    set_cell(ws2, row, 4, entry['muscle_ru'], None, None, wrap_align, thin_border)
    set_cell(ws2, row, 5, entry['head_ru'], None, None, wrap_align, thin_border)
    
    for model_col, model_name in enumerate(MODEL_NAMES, 6):
        cnt = coverage.get((cat_idx, model_name), 0)
        fill = green_fill if cnt >= 2 else (yellow_fill if cnt == 1 else red_fill)
        set_cell(ws2, row, model_col, cnt, None, fill, center_align, thin_border)
    
    in_archery = "Да" if cat_idx in archery_catalog_indices else "Нет"
    arch_col = 6 + len(MODEL_NAMES)
    fill = green_fill if in_archery == "Да" else PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    set_cell(ws2, row, arch_col, in_archery, None, fill, center_align, thin_border)

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 22
ws2.column_dimensions['D'].width = 28
ws2.column_dimensions['E'].width = 25
for i in range(6, 6 + len(MODEL_NAMES)):
    ws2.column_dimensions[get_column_letter(i)].width = 8
ws2.column_dimensions[get_column_letter(6 + len(MODEL_NAMES))].width = 16
ws2.freeze_panes = 'F2'

# ---- Sheet 3: Дефициты ----
print("  Лист 3: Дефициты...")
ws3 = wb.create_sheet("Дефициты")

set_cell(ws3, 1, 1, "ДЕФИЦИТЫ: Мышцы из перечня стрелка, отсутствующие в моделях", title_font)
ws3.merge_cells('A1:F1')

headers3 = ['Латинское название', 'Фаза', 'Контекст', 'Отдел каталога', 'Мышца каталога', 'Есть в моделях']
for col_idx, h in enumerate(headers3, 1):
    set_cell(ws3, 3, col_idx, h, header_font_white, header_fill, center_align, thin_border)

row = 4
for ln, info in sorted(latin_to_catalog.items()):
    entries = info['entries']
    phase = info['archery_info']['phase'][:60]
    context = info['archery_info']['russian_context'][:80]
    
    if not entries:
        set_cell(ws3, row, 1, ln, None, red_fill, wrap_align, thin_border)
        set_cell(ws3, row, 2, phase, None, red_fill, wrap_align, thin_border)
        set_cell(ws3, row, 3, context, None, red_fill, wrap_align, thin_border)
        reason = "НЕ В КАТАЛОГЕ" if info['method'] == 'NOT_FOUND' else "НЕ В МОДЕЛЯХ"
        set_cell(ws3, row, 4, reason, None, red_fill, center_align, thin_border)
        set_cell(ws3, row, 5, "—", None, red_fill, center_align, thin_border)
        set_cell(ws3, row, 6, "НИ В ОДНОЙ", None, red_fill, center_align, thin_border)
        row += 1
        continue
    
    # Check which models have this muscle
    models_with = []
    for entry in entries:
        try:
            cat_idx = catalog.index(entry)
        except ValueError:
            continue
        for model_name in MODEL_NAMES:
            if coverage.get((cat_idx, model_name), 0) > 0:
                if model_name not in models_with:
                    models_with.append(model_name)
    
    # Only show if not all models have it, or if it's a deficit
    if len(models_with) < len(MODEL_NAMES):
        dept_str = entries[0]['dept'] if entries else "—"
        muscle_str = entries[0]['muscle_ru'] if entries else "—"
        
        fill = red_fill if not models_with else (orange_fill if len(models_with) <= 3 else None)
        
        set_cell(ws3, row, 1, ln, None, fill, wrap_align, thin_border)
        set_cell(ws3, row, 2, phase, None, fill, wrap_align, thin_border)
        set_cell(ws3, row, 3, context, None, fill, wrap_align, thin_border)
        set_cell(ws3, row, 4, dept_str, None, fill, center_align, thin_border)
        set_cell(ws3, row, 5, muscle_str, None, fill, wrap_align, thin_border)
        set_cell(ws3, row, 6, ", ".join(models_with) if models_with else "НИ В ОДНОЙ", 
                 None, fill, center_align, thin_border)
        row += 1

ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 40
ws3.column_dimensions['C'].width = 50
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 25
ws3.column_dimensions['F'].width = 45

# ---- Sheet 4: Стратегия сборки ----
print("  Лист 4: Стратегия сборки...")
ws4 = wb.create_sheet("Стратегия сборки")

set_cell(ws4, 1, 1, "СТРАТЕГИЯ СБОРКИ МОДЕЛИ ЛУЧНИКА", title_font)
ws4.merge_cells('A1:G1')

headers4 = ['Отдел', 'Основная модель', '% покрытия', 'Доп. модель 1', 'Доп. модель 2',
            'Минимальный набор моделей', 'Комментарий']
for col_idx, h in enumerate(headers4, 1):
    set_cell(ws4, 3, col_idx, h, header_font_white, header_fill, center_align, thin_border)

row = 4
for region in regions:
    model_scores = [(mn, region_coverage[region]['by_model'][mn]['pct'],
                     region_coverage[region]['by_model'][mn]['present'],
                     region_coverage[region]['by_model'][mn]['total'])
                    for mn in MODEL_NAMES]
    model_scores.sort(key=lambda x: x[1], reverse=True)
    
    best = model_scores[0]
    second = model_scores[1] if len(model_scores) > 1 else ('—', 0, 0, 0)
    third = model_scores[2] if len(model_scores) > 2 else ('—', 0, 0, 0)
    
    # Greedy minimum set
    covered = set()
    for cat_idx, entry in region_coverage[region]['entries']:
        if coverage.get((cat_idx, best[0]), 0) > 0:
            covered.add(cat_idx)
    
    min_models = [best[0]]
    remaining = [ms for ms in model_scores[1:]]
    
    while len(covered) < region_coverage[region]['total'] and remaining:
        best_add, best_new = None, 0
        for ms in remaining:
            new = sum(1 for cat_idx, _ in region_coverage[region]['entries']
                      if cat_idx not in covered and coverage.get((cat_idx, ms[0]), 0) > 0)
            if new > best_new:
                best_new = new
                best_add = ms
        if best_add and best_new > 0:
            min_models.append(best_add[0])
            for cat_idx, _ in region_coverage[region]['entries']:
                if coverage.get((cat_idx, best_add[0]), 0) > 0:
                    covered.add(cat_idx)
            remaining.remove(best_add)
        else:
            break
    
    total_r = region_coverage[region]['total']
    covered_n = len(covered)
    min_pct = 100 * covered_n / total_r if total_r > 0 else 0
    
    set_cell(ws4, row, 1, region, header_font, None, wrap_align, thin_border)
    set_cell(ws4, row, 2, best[0], None, green_fill, center_align, thin_border)
    set_cell(ws4, row, 3, f"{best[1]:.1f}% ({best[2]}/{best[3]})", None, None, center_align, thin_border)
    set_cell(ws4, row, 4, f"{second[0]} ({second[1]:.1f}%)", None, None, center_align, thin_border)
    set_cell(ws4, row, 5, f"{third[0]} ({third[1]:.1f}%)", None, None, center_align, thin_border)
    set_cell(ws4, row, 6, " + ".join(min_models), None, None, wrap_align, thin_border)
    set_cell(ws4, row, 7, f"Мин. набор: {covered_n}/{total_r} ({min_pct:.1f}%)", None, None, wrap_align, thin_border)
    row += 1

# Global ranking
row += 2
set_cell(ws4, row, 1, "ГЛОБАЛЬНАЯ РЕКОМЕНДАЦИЯ", title_font)
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 2

overall_scores = []
for model_name in MODEL_NAMES:
    total_present = sum(region_coverage[r]['by_model'][model_name]['present'] for r in regions)
    total_expected = sum(region_coverage[r]['total'] for r in regions)
    pct = 100 * total_present / total_expected if total_expected > 0 else 0
    overall_scores.append((model_name, pct, total_present, total_expected))
overall_scores.sort(key=lambda x: x[1], reverse=True)

set_cell(ws4, row, 1, "Рейтинг моделей по общему покрытию:", Font(bold=True, size=12))
row += 1
for rank, (mn, pct, pres, tot) in enumerate(overall_scores, 1):
    fill = green_fill if rank == 1 else (yellow_fill if rank == 2 else (orange_fill if rank == 3 else None))
    set_cell(ws4, row, 1, f"{rank}. {mn}", None, fill, None, thin_border)
    set_cell(ws4, row, 2, f"{pct:.1f}%", None, fill, center_align, thin_border)
    set_cell(ws4, row, 3, f"{pres}/{tot} мышц", None, fill, center_align, thin_border)
    row += 1

# Minimum global model set
row += 2
set_cell(ws4, row, 1, "Минимальный набор моделей для максимального покрытия:", Font(bold=True, size=12))
row += 1

all_covered = set()
min_global_models = []
remaining = list(overall_scores)

while len(all_covered) < len(catalog) and remaining:
    best_add, best_new = None, 0
    for ms in remaining:
        new = sum(1 for ci in range(len(catalog)) 
                  if ci not in all_covered and coverage.get((ci, ms[0]), 0) > 0)
        if new > best_new:
            best_new = new
            best_add = ms
    if best_add and best_new > 0:
        min_global_models.append((best_add[0], best_new))
        for ci in range(len(catalog)):
            if coverage.get((ci, best_add[0]), 0) > 0:
                all_covered.add(ci)
        remaining.remove(best_add)
    else:
        break

cumulative = 0
for i, (mn, new) in enumerate(min_global_models, 1):
    cumulative += new
    fill = green_fill if i == 1 else (yellow_fill if i == 2 else (orange_fill if i == 3 else None))
    set_cell(ws4, row, 1, f"{i}. {mn}", None, fill, None, thin_border)
    set_cell(ws4, row, 2, f"+{new} новых мышц", None, fill, center_align, thin_border)
    set_cell(ws4, row, 3, f"Нарастающий итог: {cumulative}/{len(catalog)}", None, fill, center_align, thin_border)
    row += 1

total_coverable = len(all_covered)
uncoverable = len(catalog) - total_coverable
row += 1
set_cell(ws4, row, 1, f"Итого покрыто: {total_coverable}/{len(catalog)} ({100*total_coverable/len(catalog):.1f}%) мышц каталога", 
         Font(bold=True, size=11))
if uncoverable > 0:
    row += 1
    set_cell(ws4, row, 1, f"Непокрываемых мышц: {uncoverable} (отсутствуют во всех 11 моделях)", 
             Font(bold=True, size=11, color='FF0000'))
    # List uncoverable muscles
    row += 1
    for ci in range(len(catalog)):
        if ci not in all_covered:
            e = catalog[ci]
            set_cell(ws4, row, 1, f"  — {e['muscle_ru']} ({e['head_ru']}) [{e['dept']}]", small_font)
            row += 1

ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 22
ws4.column_dimensions['C'].width = 22
ws4.column_dimensions['D'].width = 20
ws4.column_dimensions['E'].width = 20
ws4.column_dimensions['F'].width = 35
ws4.column_dimensions['G'].width = 35

wb.save(OUTPUT_FILE)
print(f"\n  XLSX сохранён: {OUTPUT_FILE}")

# ============================================================
# FINAL SUMMARY
# ============================================================
print("\n" + "=" * 80)
print("ИТОГОВЫЙ ОТЧЁТ")
print("=" * 80)

print("\n1. ОБЩЕЕ ПОКРЫТИЕ МОДЕЛЕЙ (каталог: {} мышц):".format(len(catalog)))
for rank, (mn, pct, pres, tot) in enumerate(overall_scores, 1):
    marker = "★" if rank == 1 else ("☆" if rank <= 3 else " ")
    print(f"  {marker} {rank:2d}. {mn:10s} — {pct:5.1f}% ({pres}/{tot})")

print(f"\n2. ЛУЧШИЕ МОДЕЛИ ПО ОТДЕЛАМ:")
for region in regions:
    model_scores = [(mn, region_coverage[region]['by_model'][mn]['pct'],
                     region_coverage[region]['by_model'][mn]['present'],
                     region_coverage[region]['by_model'][mn]['total'])
                    for mn in MODEL_NAMES]
    model_scores.sort(key=lambda x: x[1], reverse=True)
    best = model_scores[0]
    second = model_scores[1]
    print(f"  {region:25s}: {best[0]:10s} ({best[1]:5.1f}%, {best[2]}/{best[3]}) | 2-е: {second[0]:10s} ({second[1]:5.1f}%)")

print(f"\n3. КРИТИЧЕСКИЕ ДЕФИЦИТЫ:")
print(f"  Мышц из перечня стрелка, отсутствующих во ВСЕХ моделях: {len(no_model_muscles)}")
for ln, info, reason in no_model_muscles:
    print(f"    - {ln}: {reason}")
print(f"  Мышц из перечня стрелка, только в 1 модели: {len(single_model_muscles)}")
for ln, info, model in single_model_muscles:
    print(f"    - {ln} (только {model})")

print(f"\n4. СТРАТЕГИЯ СБОРКИ (минимальный набор для максимального покрытия):")
for i, (mn, new) in enumerate(min_global_models, 1):
    print(f"    {i}. {mn} (+{new} мышц)")
print(f"  Итого покрываемых: {total_coverable}/{len(catalog)} ({100*total_coverable/len(catalog):.1f}%)")
print(f"  Непокрываемых: {uncoverable}")

print(f"\n5. СПЕЦИФИКА ПО ОТДЕЛАМ:")
for region in regions:
    model_scores = [(mn, region_coverage[region]['by_model'][mn]) for mn in MODEL_NAMES]
    model_scores.sort(key=lambda x: x[1]['pct'], reverse=True)
    top3 = model_scores[:3]
    top3_str = ", ".join(f"{mn}({rc['pct']:.0f}%)" for mn, rc in top3)
    print(f"  {region:25s}: Топ-3: {top3_str}")

print("\n" + "=" * 80)
print(f"Файл сохранён: {OUTPUT_FILE}")
print("=" * 80)
