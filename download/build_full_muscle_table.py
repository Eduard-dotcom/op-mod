import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── Load data ──
with open('/home/z/my-project/download/muscle_catalog.json', encoding='utf-8') as f:
    catalog = json.load(f)
with open('/home/z/my-project/download/all_muscles_v3.json', encoding='utf-8') as f:
    all_muscles = json.load(f)

MODELS = ['M7', 'M7_18', 'M8', 'M8_Corr', 'M8_Norm', 'M2', 'M4', 'M6', 'M9', 'Raj', 'ULB']

# ── Helper ──
def strip_side(name):
    for s in ['_r', '_l', '_R', '_L']:
        if name.endswith(s):
            return name[:-2]
    return name

# ── Build muscle lookup: model → {full_name: muscle_data} ──
muscle_lookup = {}
for model_name, muscles in all_muscles.items():
    lookup = {}
    for m in muscles:
        lookup[m['name']] = m
        # Also index by stripped name → list
        base = strip_side(m['name'])
        if base not in lookup:
            lookup[base + '___list'] = []
        lookup[base + '___list'].append(m)
    muscle_lookup[model_name] = lookup

# ── Expand catalog with missing muscles ──
# First, identify which base_names from models are not covered
catalog_bases = set()
for entry in catalog:
    for bn in entry['base_names']:
        catalog_bases.add(bn)

# Add missing entries to catalog
additional = []

# LTpT_T1-T12: Длиннейшая грудная (thoracic attachments)
for i in range(1, 13):
    bn = f'LTpT_T{i}'
    if bn not in catalog_bases:
        additional.append({
            'dept': 'Туловище',
            'subgroup': 'Выпрямитель позвоночника',
            'muscle_ru': 'Длиннейшая грудная',
            'head_ru': f'Грудной {i}',
            'base_names': [bn]
        })

# LTpT_R12
if 'LTpT_R12' not in catalog_bases:
    additional.append({
        'dept': 'Туловище',
        'subgroup': 'Выпрямитель позвоночника',
        'muscle_ru': 'Длиннейшая грудная',
        'head_ru': 'Ребро 12',
        'base_names': ['LTpT_R12']
    })

# LD (Latissimus Dorsi) heads
ld_heads = [
    ('LD_L5', 'L5-рёбра'),
    ('LD_R11', 'Рёбра 5-11'),
    ('LD_R12', 'Рёбра 5-12'),
]
for bn, head in ld_heads:
    if bn not in catalog_bases:
        additional.append({
            'dept': 'Туловище',
            'subgroup': 'Широчайшая мышца спины',
            'muscle_ru': 'Широчайшая мышца спины',
            'head_ru': head,
            'base_names': [bn]
        })

# QL missing variants
ql_heads = [
    ('QL_post_I_1-L3', 'Задняя L1-L3'),
    ('QL_post_I_2-L2', 'Задняя L2-L2'),
    ('QL_post_I_2-L3', 'Задняя L2-L3'),
    ('QL_post_I_2-L4', 'Задняя L2-L4'),
    ('QL_mid_L4-12_3', 'Средняя L4-12 (3)'),
]
for bn, head in ql_heads:
    if bn not in catalog_bases:
        additional.append({
            'dept': 'Туловище',
            'subgroup': 'Квадратная мышца поясницы',
            'muscle_ru': 'Квадратная мышца поясницы',
            'head_ru': head,
            'base_names': [bn]
        })

# Intercostal IS2-8 part 7 (additional segments)
for is_num in range(2, 9):
    for prefix, ru_name in [('ExtIC', 'Наружные межрёберные'), ('IntIC', 'Внутренние межрёберные')]:
        bn = f'{prefix}_IS{is_num}_7'
        if bn not in catalog_bases:
            additional.append({
                'dept': 'Туловище',
                'subgroup': 'Межрёберные',
                'muscle_ru': ru_name,
                'head_ru': f'IS{is_num} ч.7',
                'base_names': [bn]
            })

# LAT1/2/3 (latissimus dorsi in M8/ULB)
for i in range(1, 4):
    bn = f'LAT{i}'
    if bn not in catalog_bases:
        additional.append({
            'dept': 'Туловище',
            'subgroup': 'Широчайшая мышца спины',
            'muscle_ru': 'Широчайшая мышца спины',
            'head_ru': f'Головка {i}',
            'base_names': [bn]
        })

# ercspn (erector spinae in M8/ULB)
if 'ercspn' not in catalog_bases:
    additional.append({
        'dept': 'Туловище',
        'subgroup': 'Выпрямитель позвоночника',
        'muscle_ru': 'Выпрямитель позвоночника',
        'head_ru': 'Единая мышца',
        'base_names': ['ercspn']
    })

# extobl / intobl (oblique in ULB)
for bn, ru, head in [('extobl', 'Наружная косая живота', 'Единая мышца'), ('intobl', 'Внутренняя косая живота', 'Единая мышца')]:
    if bn not in catalog_bases:
        additional.append({
            'dept': 'Туловище',
            'subgroup': 'Косые мышцы живота',
            'muscle_ru': ru,
            'head_ru': head,
            'base_names': [bn]
        })

# SerrAnt2_2 (serratus anterior)
if 'SerrAnt2_2' not in catalog_bases:
    additional.append({
        'dept': 'Туловище',
        'subgroup': 'Передняя зубчатая',
        'muscle_ru': 'Передняя зубчатая',
        'head_ru': 'Часть 2',
        'base_names': ['SerrAnt2_2']
    })

# multifidus T3-T6
for t in range(3, 7):
    bn = f'multifidus_T{t}_T{t-2}'
    if bn not in catalog_bases:
        additional.append({
            'dept': 'Туловище',
            'subgroup': 'Многораздельная',
            'muscle_ru': 'Многораздельная',
            'head_ru': f'T{t}-T{t-2}',
            'base_names': [bn]
        })

# splen_cap_sklc6 (M6 specific)
if 'splen_cap_sklc6' not in catalog_bases:
    additional.append({
        'dept': 'Шея',
        'subgroup': 'Ременные',
        'muscle_ru': 'Ременная мышца головы',
        'head_ru': 'skull c6',
        'base_names': ['splen_cap_sklc6']
    })

catalog_extended = catalog + additional

# ── Recheck coverage ──
catalog_bases_ext = set()
for entry in catalog_extended:
    for bn in entry['base_names']:
        catalog_bases_ext.add(bn)

unmatched_count = 0
for model, muscles in all_muscles.items():
    for m in muscles:
        base = strip_side(m['name'])
        if base not in catalog_bases_ext and base not in ['default', 'defaultmuscle']:
            unmatched_count += 1

print(f"Extended catalog: {len(catalog_extended)} entries, {len(catalog_bases_ext)} base names")
print(f"Remaining unmatched muscles: {unmatched_count}")

# ── Build workbook ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Мышцы'

# Styles
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
sub_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
data_font = Font(name='Arial', size=9)
num_font = Font(name='Arial', size=9, bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Row 1: Main headers
fixed_headers = ['№', 'Отдел', 'Подгруппа', 'Мышца', 'Головка']
for i, h in enumerate(fixed_headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border

col = 6
for model in MODELS:
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)
    c = ws.cell(row=1, column=col, value=model)
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border
    for dc in range(3):
        ws.cell(row=1, column=col+dc).border = thin_border
    col += 3

# Row 2: Sub-headers
for i in range(1, 6):
    c = ws.cell(row=2, column=i, value='')
    c.fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    c.border = thin_border

col = 6
for model in MODELS:
    for j, sh in enumerate(['Путь', 'Начало', 'Конец']):
        c = ws.cell(row=2, column=col+j, value=sh)
        c.font = sub_font; c.fill = sub_header_fill; c.alignment = center_align; c.border = thin_border
    col += 3

# Data rows
row = 3
num = 1

for entry in catalog_extended:
    dept = entry['dept']
    subgroup = entry['subgroup']
    muscle_ru = entry['muscle_ru']
    head_ru = entry['head_ru']
    base_names = entry['base_names']

    # A: №
    ws.cell(row=row, column=1, value=num).font = num_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='top')
    ws.cell(row=row, column=1).border = thin_border

    # B: Отдел — ВСЕГДА заполнен
    ws.cell(row=row, column=2, value=dept).font = data_font
    ws.cell(row=row, column=2).alignment = wrap_align
    ws.cell(row=row, column=2).border = thin_border

    # C: Подгруппа — ВСЕГДА заполнена
    ws.cell(row=row, column=3, value=subgroup).font = data_font
    ws.cell(row=row, column=3).alignment = wrap_align
    ws.cell(row=row, column=3).border = thin_border

    # D: Мышца — ВСЕГДА заполнена
    ws.cell(row=row, column=4, value=muscle_ru).font = data_font
    ws.cell(row=row, column=4).alignment = wrap_align
    ws.cell(row=row, column=4).border = thin_border

    # E: Головка — ВСЕГДА заполнена
    ws.cell(row=row, column=5, value=head_ru).font = data_font
    ws.cell(row=row, column=5).alignment = wrap_align
    ws.cell(row=row, column=5).border = thin_border

    # Model columns
    col = 6
    for model in MODELS:
        found_muscle = None
        lookup = muscle_lookup.get(model, {})

        for bn in base_names:
            # Try exact names first (with _r/_l suffixes)
            for suffix in ['_r', '_l', '_R', '_L', '']:
                key = bn + suffix
                if key in lookup and not key.endswith('___list'):
                    found_muscle = lookup[key]
                    break
                # Try base name directly
                list_key = bn + '___list'
                if list_key in lookup:
                    # Pick the first matching muscle
                    found_muscle = lookup[list_key][0]
                    break
            if found_muscle:
                break

        if found_muscle:
            bodies = found_muscle['bodies']
            path_str = ' → '.join(bodies) if bodies else ''
            origin = bodies[0] if bodies else ''
            insertion = bodies[-1] if len(bodies) > 1 else (bodies[0] if bodies else '')

            path_cell = f"{found_muscle['name']}\n{path_str}" if path_str else found_muscle['name']
            ws.cell(row=row, column=col, value=path_cell).font = data_font
            ws.cell(row=row, column=col).alignment = wrap_align
            ws.cell(row=row, column=col).border = thin_border

            ws.cell(row=row, column=col+1, value=origin).font = data_font
            ws.cell(row=row, column=col+1).alignment = wrap_align
            ws.cell(row=row, column=col+1).border = thin_border

            ws.cell(row=row, column=col+2, value=insertion).font = data_font
            ws.cell(row=row, column=col+2).alignment = wrap_align
            ws.cell(row=row, column=col+2).border = thin_border
        else:
            for dc in range(3):
                ws.cell(row=row, column=col+dc, value='—').font = data_font
                ws.cell(row=row, column=col+dc).alignment = Alignment(horizontal='center', vertical='top')
                ws.cell(row=row, column=col+dc).border = thin_border
        col += 3

    row += 1
    num += 1

# Column widths
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 24
ws.column_dimensions['E'].width = 18
for i in range(len(MODELS)):
    base_col = 6 + i * 3
    ws.column_dimensions[get_column_letter(base_col)].width = 32
    ws.column_dimensions[get_column_letter(base_col+1)].width = 14
    ws.column_dimensions[get_column_letter(base_col+2)].width = 14

ws.freeze_panes = 'F3'

# Legend sheet
ws2 = wb.create_sheet('Как читать таблицу')
legend = [
    ['Элемент', 'Описание'],
    ['Столбец «Путь»', 'Имя мышцы в модели (1-я строка) + полный путь через тела (2-я строка): тело1 → тело2 → ... → телоN'],
    ['Столбец «Начало»', 'Первое тело в пути мышцы (origin / точка начала)'],
    ['Столбец «Конец»', 'Последнее тело в пути мышцы (insertion / точка вставки)'],
    ['—', 'Мышца отсутствует в данной модели'],
    ['Головка', 'Для многоглавых мышц — название конкретной головки; для одноглавых — «Единая мышца»'],
    ['Все ячейки заполнены', 'Каждая ячейка содержит значение для корректной сортировки/фильтрации'],
    ['M4', 'Модель без мышц (ScapulothoracicJoint_Shoulder)'],
]
for r, row_data in enumerate(legend, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = Font(name='Arial', size=10, bold=(r==1))
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 80

out_path = '/home/z/my-project/download/Мышцы_все_модели.xlsx'
wb.save(out_path)
print(f'Saved: {out_path}')
print(f'Rows: {num-1}, Models: {MODELS}')
