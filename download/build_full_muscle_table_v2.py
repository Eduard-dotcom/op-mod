import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── Load data ──
with open('/home/z/my-project/download/all_muscles_v3.json', encoding='utf-8') as f:
    all_muscles = json.load(f)

MODELS = ['M7', 'M7_18', 'M8', 'M8_Corr', 'M8_Norm', 'M2', 'M4', 'M6', 'M9', 'Raj', 'ULB']

def strip_side(name):
    for s in ['_r', '_l', '_R', '_L']:
        if name.endswith(s):
            return name[:-2]
    return name

def get_side(name):
    if name.endswith(('_r', '_R')):
        return 'R'
    elif name.endswith(('_l', '_L')):
        return 'L'
    return 'Обе'

SKIP_NAMES = {'default', 'defaultmuscle'}

# ── Build anatomical catalog from ALL models' muscles ──
# Group: (base_name) → list of muscles across all models
all_by_base = {}
for model, muscles in all_muscles.items():
    for m in muscles:
        base = strip_side(m['name'])
        if base in SKIP_NAMES:
            continue
        if base not in all_by_base:
            all_by_base[base] = {}
        if model not in all_by_base[base]:
            all_by_base[base][model] = []
        all_by_base[base][model].append(m)

print(f"Уникальных базовых имён мышц: {len(all_by_base)}")

# ── Anatomical classification ──
# Build from the existing catalog + automatic classification for unknowns
with open('/home/z/my-project/download/muscle_catalog.json', encoding='utf-8') as f:
    orig_catalog = json.load(f)

# Map base_name → anatomy from existing catalog
anatomy_map = {}  # base_name → {dept, subgroup, muscle_ru, head_ru}
for entry in orig_catalog:
    for bn in entry['base_names']:
        anatomy_map[bn] = {
            'dept': entry['dept'],
            'subgroup': entry['subgroup'],
            'muscle_ru': entry['muscle_ru'],
            'head_ru': entry['head_ru']
        }

# Auto-classify unknown base names based on prefix patterns
def auto_classify(base):
    """Classify muscle by its base name pattern"""
    # Psoas
    if base.startswith('Ps_') or base.startswith('psoas'):
        if 'VB' in base:
            lvl = base.split('L')[1].split('_')[0] if 'L' in base else '?'
            return 'Туловище', 'Подвздошно-поясничная', 'Подвздошно-поясничная', f'Тело позвонка L{lvl}'
        elif 'TP' in base:
            lvl = base.split('L')[1].split('_')[0] if 'L' in base else '?'
            return 'Туловище', 'Подвздошно-поясничная', 'Подвздошно-поясничная', f'Поперечный отросток L{lvl}'
        elif 'IVD' in base:
            return 'Туловище', 'Подвздошно-поясничная', 'Подвздошно-поясничная', 'Межпозв. диск'
        else:
            return 'Туловище', 'Подвздошно-поясничная', 'Подвздошно-поясничная', 'Единая мышца'
    
    if base == 'iliacus':
        return 'Туловище', 'Подвздошно-поясничная', 'Подвздошная', 'Единая мышца'
    
    # Erector spinae - Longissimus Thoracis (LTpT)
    if base.startswith('LTpT_T'):
        lvl = base.split('T')[1].split('_')[0]
        return 'Туловище', 'Выпрямитель позвоночника', 'Длиннейшая грудная', f'Грудной T{lvl}'
    if base.startswith('LTpT_R'):
        lvl = base.split('R')[1].split('_')[0]
        return 'Туловище', 'Выпрямитель позвоночника', 'Длиннейшая грудная', f'Ребро {lvl}'
    if base.startswith('LTpL_L'):
        lvl = base.split('L')[1].split('_')[0]
        return 'Туловище', 'Выпрямитель позвоночника', 'Длиннейшая поясничная', f'L{lvl}'
    
    # Iliocostalis (IL_)
    if base.startswith('IL_R'):
        lvl = base.split('R')[1]
        return 'Туловище', 'Выпрямитель позвоночника', 'Подвздошно-рёберная', f'Ребро {lvl}'
    if base.startswith('IL_L'):
        lvl = base.split('L')[1]
        return 'Туловище', 'Выпрямитель позвоночника', 'Подвздошно-рёберная', f'L{lvl}'
    if base.startswith('iliocost_cerv'):
        return 'Туловище', 'Выпрямитель позвоночника', 'Подвздошно-рёберная шеи', 'C5-ребро'
    
    # Latissimus Dorsi (LD_)
    if base.startswith('LD_'):
        return 'Туловище', 'Широчайшая мышца спины', 'Широчайшая мышца спины', base.replace('LD_', '')
    
    # LAT (latissimus in M8/ULB)
    if base.startswith('LAT'):
        return 'Туловище', 'Широчайшая мышца спины', 'Широчайшая мышца спины', f'Головка {base[-1]}'
    
    # Erector spinae (ercspn)
    if base == 'ercspn':
        return 'Туловище', 'Выпрямитель позвоночника', 'Выпрямитель позвоночника', 'Единая мышца'
    
    # Quadratus Lumborum (QL_)
    if base.startswith('QL_ant'):
        return 'Туловище', 'Квадратная мышца поясницы', 'Квадратная мышца поясницы', f'Передняя {base[7:]}'
    if base.startswith('QL_mid'):
        return 'Туловище', 'Квадратная мышца поясницы', 'Квадратная мышца поясницы', f'Средняя {base[7:]}'
    if base.startswith('QL_post'):
        return 'Туловище', 'Квадратная мышца поясницы', 'Квадратная мышца поясницы', f'Задняя {base[8:]}'
    
    # Multifidus (MF_ / multifidus_)
    if base.startswith('MF_') or base.startswith('multifidus_'):
        if base.startswith('MF_'):
            tag = base[3:]
        else:
            tag = base[11:]
        return 'Туловище', 'Многораздельная', 'Многораздельная', tag
    
    # Intercostal (ExtIC_ / IntIC_)
    if base.startswith('ExtIC_'):
        tag = base[6:]
        return 'Туловище', 'Межрёберные', 'Наружные межрёберные', tag
    if base.startswith('IntIC_'):
        tag = base[6:]
        return 'Туловище', 'Межрёберные', 'Внутренние межрёберные', tag
    
    # Obliques
    if base in ['extobl', 'EO']:
        return 'Туловище', 'Косые мышцы живота', 'Наружная косая живота', 'Единая мышца'
    if base in ['intobl', 'IO']:
        return 'Туловище', 'Косые мышцы живота', 'Внутренняя косая живота', 'Единая мышца'
    if base.startswith('EO'):
        return 'Туловище', 'Косые мышцы живота', 'Наружная косая живота', f'Часть {base[2:]}'
    if base.startswith('IO'):
        return 'Туловище', 'Косые мышцы живота', 'Внутренняя косая живота', f'Часть {base[2:]}'
    
    # Rectus abdominis
    if base.startswith('rect_abd'):
        return 'Туловище', 'Прямая мышца живота', 'Прямая мышца живота', 'Единая мышца'
    
    # Serratus anterior
    if base.startswith('SerrAnt'):
        return 'Туловище', 'Передняя зубчатая', 'Передняя зубчатая', f'Часть {base[7:]}'
    
    # ---- NECK ----
    # Sternocleidomastoid
    if base in ['stern_mast']:
        return 'Шея', 'Сгибатели', 'Грудино-ключично-сосцевидная', 'Единая мышца'
    if base in ['cleid_mast']:
        return 'Шея', 'Сгибатели', 'Ключично-сосцевидная', 'Единая мышца'
    if base in ['cleid_occ']:
        return 'Шея', 'Сгибатели', 'Ключично-затылочная', 'Единая мышца'
    
    # Scalenes
    if base == 'scalenus_ant':
        return 'Шея', 'Лестничные', 'Передняя лестничная', 'Единая мышца'
    if base == 'scalenus_med':
        return 'Шея', 'Лестничные', 'Средняя лестничная', 'Единая мышца'
    if base == 'scalenus_post':
        return 'Шея', 'Лестничные', 'Задняя лестничная', 'Единая мышца'
    
    # Longus
    if base.startswith('long_cap'):
        return 'Шея', 'Длинные мышцы', 'Длинная мышца головы', 'skull'
    if base.startswith('long_col_'):
        return 'Шея', 'Длинные мышцы', 'Длинная мышца шеи', base[9:]
    if base.startswith('longissi_cap'):
        return 'Шея', 'Длинные мышцы', 'Длиннейшая мышцы головы', 'skull'
    if base.startswith('longissi_cerv'):
        return 'Шея', 'Длинные мышцы', 'Длиннейшая мышцы шеи', base[13:]
    
    # Suboccipital
    if base == 'rectcap_post_maj':
        return 'Шея', 'Подзатылочные', 'Задняя большая прямая мышца головы', 'Единая мышца'
    if base == 'rectcap_post_min':
        return 'Шея', 'Подзатылочные', 'Задняя малая прямая мышца головы', 'Единая мышца'
    if base == 'obl_cap_inf':
        return 'Шея', 'Подзатылочные', 'Нижняя косая мышца головы', 'Единая мышца'
    if base == 'obl_cap_sup':
        return 'Шея', 'Подзатылочные', 'Верхняя косая мышца головы', 'Единая мышца'
    
    # Splenius
    if base.startswith('semi_cap') or base.startswith('splen_cap'):
        return 'Шея', 'Ременные', 'Ременная мышца головы', base.split('_', 2)[-1] if '_' in base else 'Единая мышца'
    if base.startswith('semi_cerv') or base.startswith('splen_cerv'):
        return 'Шея', 'Ременные', 'Ременная мышца шеи', base.split('_', 2)[-1] if '_' in base else 'Единая мышца'
    
    # Deep/superior multifidus (cervical)
    if base.startswith('deepmult'):
        return 'Шея', 'Многораздельные', 'Глубокая многораздельная', base[9:]
    if base.startswith('supmult'):
        return 'Шея', 'Многораздельные', 'Верхняя многораздельная', base[8:]
    
    # Infrahyoid
    if base == 'sternohyoid':
        return 'Шея', 'Подъязычные', 'Грудино-подъязычная', 'Единая мышца'
    if base == 'sternothyroid':
        return 'Шея', 'Подъязычные', 'Грудино-щитовидная', 'Единая мышца'
    if base == 'omohyoid':
        return 'Шея', 'Подъязычные', 'Лопаточно-подъязычная', 'Единая мышца'
    
    # ---- UPPER LIMB ----
    # Deltoid
    if base.startswith('DELT') or base.startswith('deltoid'):
        return 'Верхняя конечность', 'Дельтовидная', 'Дельтовидная', f'Головка {base[-1]}' if base[-1].isdigit() else 'Единая мышца'
    
    # Trapezius
    if base.startswith('TRAP') or base.startswith('trap'):
        return 'Верхняя конечность', 'Трапециевидная', 'Трапециевидная', f'Часть {base[-1]}' if base[-1].isdigit() else 'Единая мышца'
    
    # Rotator cuff
    if base in ['supraspinatus', 'SUPSP']:
        return 'Верхняя конечность', 'Вращательная манжета', 'Надостная', 'Единая мышца'
    if base in ['infraspinatus', 'INFSP']:
        return 'Верхняя конечность', 'Вращательная манжета', 'Подостная', 'Единая мышца'
    if base in ['subscapularis', 'SUBSC']:
        return 'Верхняя конечность', 'Вращательная манжета', 'Подлопаточная', 'Единая мышца'
    if base in ['teres_minor', 'TMIN']:
        return 'Верхняя конечность', 'Вращательная манжета', 'Малая круглая', 'Единая мышца'
    
    # Biceps
    if base.startswith('BIC') or base.startswith('biceps'):
        return 'Верхняя конечность', 'Двуглавая плеча', 'Двуглавая плеча', 'Единая мышца'
    
    # Triceps
    if base.startswith('TRI') or base.startswith('triceps'):
        if 'long' in base.lower():
            return 'Верхняя конечность', 'Трёхглавая плеча', 'Трёхглавая плеча', 'Длинная головка'
        elif 'lat' in base.lower():
            return 'Верхняя конечность', 'Трёхглавая плеча', 'Трёхглавая плеча', 'Латеральная головка'
        elif 'med' in base.lower():
            return 'Верхняя конечность', 'Трёхглавая плеча', 'Трёхглавая плеча', 'Медиальная головка'
        return 'Верхняя конечность', 'Трёхглавая плеча', 'Трёхглавая плеча', 'Единая мышца'
    
    # Forearm/wrist/hand - classify broadly
    if any(kw in base.lower() for kw in ['flex', 'flexor', 'flexcar', 'flexdig', 'flexpol', 'FCR', 'FCU', 'FDS', 'FDP', 'FPL']):
        return 'Верхняя конечность', 'Сгибатели предплечья', 'Сгибатель', base
    if any(kw in base.lower() for kw in ['ext', 'extensor', 'extcar', 'extdig', 'extpol', 'ECRB', 'ECRL', 'ECU', 'EDC', 'EIP', 'EPL']):
        return 'Верхняя конечность', 'Разгибатели предплечья', 'Разгибатель', base
    if any(kw in base.lower() for kw in ['pron', 'pronator', 'supin', 'supinator', 'brach', 'brachiorad', 'BRD', 'PT', 'PQ']):
        return 'Верхняя конечность', 'Пронаторы/супинаторы', 'Пронатор/супинатор', base
    
    # ---- LOWER LIMB ----
    if any(kw in base.lower() for kw in ['glut', 'gluteus', 'GLM', 'GMAX', 'GMED', 'GMIN']):
        return 'Нижняя конечность', 'Ягодичные', 'Ягодичная', base
    if any(kw in base.lower() for kw in ['rect_fem', 'rectus_fem', 'RF', 'vas', 'vastus', 'VASMED', 'VASLAT', 'VASINT']):
        return 'Нижняя конечность', 'Четырёхглавая бедра', 'Четырёхглавая бедра', base
    if any(kw in base.lower() for kw in ['hamstr', 'bifem', 'semi', 'semiten', 'semimem', 'BF', 'ST', 'SM']):
        if 'semi_cerv' in base or 'semi_cap' in base:
            pass  # not hamstring
        else:
            return 'Нижняя конечность', 'Задняя поверхность бедра', 'Задняя поверхность бедра', base
    if any(kw in base.lower() for kw in ['gastroc', 'gas', 'soleus', 'SOL', 'tib_post', 'tib_ant', 'TP', 'TA']):
        return 'Нижняя конечность', 'Голень', 'Мышца голени', base
    if any(kw in base.lower() for kw in ['adduct', 'ADD', 'add_mag', 'add_long', 'add_brev', 'grac', 'GRAC']):
        return 'Нижняя конечность', 'Приводящие', 'Приводящая', base
    if any(kw in base.lower() for kw in ['peri', 'peroneus', 'per_brev', 'per_long', 'PER']):
        return 'Нижняя конечность', 'Малоберцовые', 'Малоберцовая', base
    
    # Default: unknown
    return '???', '???', base, base

# Classify all base names
catalog_full = []
classified = set()
for base in sorted(all_by_base.keys()):
    if base in anatomy_map:
        a = anatomy_map[base]
        dept, subgroup, muscle_ru, head_ru = a['dept'], a['subgroup'], a['muscle_ru'], a['head_ru']
    else:
        dept, subgroup, muscle_ru, head_ru = auto_classify(base)
    
    catalog_full.append({
        'base_name': base,
        'dept': dept,
        'subgroup': subgroup,
        'muscle_ru': muscle_ru,
        'head_ru': head_ru,
    })
    classified.add(base)

# Count how many fell to default
unknown = [c for c in catalog_full if c['dept'] == '???']
print(f"Не классифицировано: {len(unknown)}")
if unknown:
    for u in unknown[:20]:
        print(f"  {u['base_name']}: models={list(all_by_base[u['base_name']].keys())}")

# ── Now determine which bases have R/L distinction ──
# A base_name has sides if ANY model has both _r and _l variants
bases_with_sides = set()
bases_without_sides = set()

for base, models in all_by_base.items():
    has_r = False
    has_l = False
    has_noside = False
    for model, ms in models.items():
        for m in ms:
            side = get_side(m['name'])
            if side == 'R': has_r = True
            elif side == 'L': has_l = True
            else: has_noside = True
    
    if has_r or has_l:
        bases_with_sides.add(base)
    else:
        bases_without_sides.add(base)

print(f"\nС R/L стороной: {len(bases_with_sides)}")
print(f"Без стороны: {len(bases_without_sides)}")

# ── Build row list: each row = (base_name, side) ──
rows = []
for entry in catalog_full:
    base = entry['base_name']
    if base in bases_with_sides:
        rows.append({**entry, 'side': 'R'})
        rows.append({**entry, 'side': 'L'})
    else:
        rows.append({**entry, 'side': 'Обе'})

print(f"\nИтого строк в таблице: {len(rows)}")

# ── Build lookup: model → {(base_name, side): muscle} ──
muscle_lookup = {}
for model, muscles in all_muscles.items():
    lookup = {}
    for m in muscles:
        base = strip_side(m['name'])
        side = get_side(m['name'])
        if base in SKIP_NAMES:
            continue
        key = (base, side)
        if key not in lookup:
            lookup[key] = m
    muscle_lookup[model] = lookup

# ── Build workbook ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Мышцы'

# Styles
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
sub_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
data_font = Font(name='Arial', size=9)
num_font = Font(name='Arial', size=9, bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Row 1: Main headers
# A=№, B=Отдел, C=Подгруппа, D=Мышца, E=Головка, F=Сторона, G+=models
fixed_headers = ['№', 'Отдел', 'Подгруппа', 'Мышца', 'Головка', 'Сторона']
for i, h in enumerate(fixed_headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border

col = 7
for model in MODELS:
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)
    c = ws.cell(row=1, column=col, value=model)
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border
    for dc in range(3):
        ws.cell(row=1, column=col+dc).border = thin_border
    col += 3

# Row 2: Sub-headers
for i in range(1, 7):
    c = ws.cell(row=2, column=i, value='')
    c.fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    c.border = thin_border

col = 7
for model in MODELS:
    for j, sh in enumerate(['Путь', 'Начало', 'Конец']):
        c = ws.cell(row=2, column=col+j, value=sh)
        c.font = sub_font; c.fill = sub_header_fill; c.alignment = center_align; c.border = thin_border
    col += 3

# Data rows
excel_row = 3
num = 1

for entry in rows:
    base = entry['base_name']
    dept = entry['dept']
    subgroup = entry['subgroup']
    muscle_ru = entry['muscle_ru']
    head_ru = entry['head_ru']
    side = entry['side']

    # A: №
    ws.cell(row=excel_row, column=1, value=num).font = num_font
    ws.cell(row=excel_row, column=1).alignment = Alignment(horizontal='center', vertical='top')
    ws.cell(row=excel_row, column=1).border = thin_border

    # B-F: always filled
    for col_idx, val in [(2, dept), (3, subgroup), (4, muscle_ru), (5, head_ru), (6, side)]:
        ws.cell(row=excel_row, column=col_idx, value=val).font = data_font
        ws.cell(row=excel_row, column=col_idx).alignment = wrap_align
        ws.cell(row=excel_row, column=col_idx).border = thin_border

    # Model columns
    col = 7
    for model in MODELS:
        lookup = muscle_lookup.get(model, {})
        
        # Try (base, R) or (base, L) or (base, Обе)
        found_muscle = lookup.get((base, side))
        if not found_muscle and side in ('R', 'L'):
            # Some models use different naming: try the other mapping
            found_muscle = lookup.get((base, 'Обе'))
        
        if found_muscle:
            bodies = found_muscle['bodies']
            path_str = ' → '.join(bodies) if bodies else ''
            origin = bodies[0] if bodies else ''
            insertion = bodies[-1] if bodies else ''

            path_cell = f"{found_muscle['name']}\n{path_str}" if path_str else found_muscle['name']
            ws.cell(row=excel_row, column=col, value=path_cell).font = data_font
            ws.cell(row=excel_row, column=col).alignment = wrap_align
            ws.cell(row=excel_row, column=col).border = thin_border

            ws.cell(row=excel_row, column=col+1, value=origin).font = data_font
            ws.cell(row=excel_row, column=col+1).alignment = wrap_align
            ws.cell(row=excel_row, column=col+1).border = thin_border

            ws.cell(row=excel_row, column=col+2, value=insertion).font = data_font
            ws.cell(row=excel_row, column=col+2).alignment = wrap_align
            ws.cell(row=excel_row, column=col+2).border = thin_border
        else:
            for dc in range(3):
                ws.cell(row=excel_row, column=col+dc, value='—').font = data_font
                ws.cell(row=excel_row, column=col+dc).alignment = Alignment(horizontal='center', vertical='top')
                ws.cell(row=excel_row, column=col+dc).border = thin_border
        col += 3

    excel_row += 1
    num += 1

# Column widths
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 26
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 8
for i in range(len(MODELS)):
    base_col = 7 + i * 3
    ws.column_dimensions[get_column_letter(base_col)].width = 34
    ws.column_dimensions[get_column_letter(base_col+1)].width = 14
    ws.column_dimensions[get_column_letter(base_col+2)].width = 14

ws.freeze_panes = 'G3'

# Legend sheet
ws2 = wb.create_sheet('Как читать таблицу')
legend = [
    ['Элемент', 'Описание'],
    ['Столбец «Путь»', 'Имя мышцы в модели (1-я строка) + полный путь через тела (2-я строка): тело1 → тело2 → ... → телоN'],
    ['Столбец «Начало»', 'Первое тело в пути мышцы (origin / точка начала)'],
    ['Столбец «Конец»', 'Последнее тело в пути мышцы (insertion / точка вставки)'],
    ['Столбец «Сторона»', 'R = правая, L = левая, Обе = модель не различает стороны'],
    ['—', 'Мышца отсутствует в данной модели на данной стороне'],
    ['Головка', 'Название конкретной головки для многоглавых мышц'],
    ['Все ячейки заполнены', 'Для корректной сортировки и фильтрации'],
    ['M4', 'Модель без мышц (ScapulothoracicJoint_Shoulder), все ячейки «—»'],
    ['Строка R и L', 'Если хотя бы в одной модели мышца имеет R/L варианты — создаются обе строки'],
]
for r, row_data in enumerate(legend, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = Font(name='Arial', size=10, bold=(r==1))
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 80

out_path = '/home/z/my-project/download/Мышцы_все_модели.xlsx'
wb.save(out_path)
print(f'\nSaved: {out_path}')
print(f'Rows: {num-1}, Models: {len(MODELS)}')
