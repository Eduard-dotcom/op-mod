import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── Load data ──
with open('/home/z/my-project/download/muscle_catalog.json', encoding='utf-8') as f:
    catalog = json.load(f)

with open('/home/z/my-project/download/all_muscles_v2.json', encoding='utf-8') as f:
    all_muscles = json.load(f)

MODELS = ['M7', 'M8', 'Raj', 'ULB']

# ── Build muscle lookup ──
def strip_side(name):
    for s in ['_r', '_l', '_R', '_L']:
        if name.endswith(s):
            return name[:-2]
    return name

muscle_lookup = {}
for model_name, muscles in all_muscles.items():
    lookup = {}
    for m in muscles:
        key = strip_side(m['name'])
        if key not in lookup:
            lookup[key] = []
        lookup[key].append(m)
    muscle_lookup[model_name] = lookup

trial_rows = catalog[:25]

# ── Create workbook ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Мышцы'

# ── Styles ──
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
sub_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
data_font = Font(name='Arial', size=9)
num_font = Font(name='Arial', size=9, bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ── Row 1: Main headers ──
fixed_headers = ['№', 'Отдел', 'Подгруппа', 'Мышца', 'Головка']
for i, h in enumerate(fixed_headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border

col = 6
for model in MODELS:
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)
    c = ws.cell(row=1, column=col, value=model)
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border
    for dc in range(3):
        ws.cell(row=1, column=col+dc).border = thin_border
    col += 3

# ── Row 2: Sub-headers ──
for i in range(1, 6):
    c = ws.cell(row=2, column=i, value='')
    c.fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    c.border = thin_border

col = 6
for model in MODELS:
    for j, sh in enumerate(['Путь', 'Начало', 'Конец']):
        c = ws.cell(row=2, column=col+j, value=sh)
        c.font = sub_font; c.fill = sub_header_fill; c.alignment = center_align; c.border = thin_border
    col += 3

# ── Data rows — EVERY cell filled, no blanks ──
row = 3
num = 1

for entry in trial_rows:
    dept = entry['dept']
    subgroup = entry['subgroup']
    muscle_ru = entry['muscle_ru']
    head_ru = entry['head_ru']
    base_names = entry['base_names']

    # A: №
    ws.cell(row=row, column=1, value=num).font = num_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='top')
    ws.cell(row=row, column=1).border = thin_border

    # B: Отдел — ВСЕГДА заполняется
    ws.cell(row=row, column=2, value=dept).font = data_font
    ws.cell(row=row, column=2).alignment = wrap_align
    ws.cell(row=row, column=2).border = thin_border

    # C: Подгруппа — ВСЕГДА заполняется
    ws.cell(row=row, column=3, value=subgroup).font = data_font
    ws.cell(row=row, column=3).alignment = wrap_align
    ws.cell(row=row, column=3).border = thin_border

    # D: Мышца — ВСЕГДА заполняется
    ws.cell(row=row, column=4, value=muscle_ru).font = data_font
    ws.cell(row=row, column=4).alignment = wrap_align
    ws.cell(row=row, column=4).border = thin_border

    # E: Головка — ВСЕГДА заполняется
    ws.cell(row=row, column=5, value=head_ru).font = data_font
    ws.cell(row=row, column=5).alignment = wrap_align
    ws.cell(row=row, column=5).border = thin_border

    # Model columns
    col = 6
    for model in MODELS:
        found_muscles = []
        lookup = muscle_lookup.get(model, {})
        for bn in base_names:
            for suffix in ['', '_r', '_l', '_R', '_L']:
                key = bn + suffix
                if key in lookup:
                    found_muscles.extend(lookup[key])

        if found_muscles:
            m = found_muscles[0]
            bodies = m['bodies']
            path_str = ' → '.join(bodies)
            origin = bodies[0] if bodies else ''
            insertion = bodies[-1] if len(bodies) > 1 else ''

            path_cell = f"{m['name']}\n{path_str}"
            ws.cell(row=row, column=col, value=path_cell).font = data_font
            ws.cell(row=row, column=col).alignment = wrap_align
            ws.cell(row=row, column=col).border = thin_border

            ws.cell(row=row, column=col+1, value=origin).font = data_font
            ws.cell(row=row, column=col+1).alignment = wrap_align
            ws.cell(row=row, column=col+1).border = thin_border

            ws.cell(row=row, column=col+2, value=insertion).font = data_font
            ws.cell(row=row, column=col+2).alignment = wrap_align
            ws.cell(row=row, column=col+2).border = thin_border
        else:
            for dc in range(3):
                ws.cell(row=row, column=col+dc, value='—').font = data_font
                ws.cell(row=row, column=col+dc).alignment = Alignment(horizontal='center', vertical='top')
                ws.cell(row=row, column=col+dc).border = thin_border
        col += 3

    row += 1
    num += 1

# ── Column widths ──
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 18
for i in range(len(MODELS)):
    base_col = 6 + i * 3
    ws.column_dimensions[get_column_letter(base_col)].width = 30
    ws.column_dimensions[get_column_letter(base_col+1)].width = 14
    ws.column_dimensions[get_column_letter(base_col+2)].width = 14

ws.freeze_panes = 'F3'

# ── Sheet 2: Legend ──
ws2 = wb.create_sheet('Как читать таблицу')
legend = [
    ['Элемент', 'Описание'],
    ['Столбец «Путь»', 'Имя мышцы в модели (1-я строка) + полный путь через тела (2-я строка): тело1 → тело2 → ... → телоN'],
    ['Столбец «Начало»', 'Первое тело в пути мышцы (origin / точка начала)'],
    ['Столбец «Конец»', 'Последнее тело в пути мышцы (insertion / точка вставки)'],
    ['—', 'Мышца отсутствует в данной модели'],
    ['Головка', 'Для многоглавых мышц — название конкретной головки; для одноглавых — «Единая мышца»'],
    ['D «Мышца»', 'Русское анатомическое название; повторяется для каждой головки'],
    ['Все ячейки заполнены', 'Каждая ячейка содержит значение — нет пустых строк для корректной сортировки и фильтрации'],
]
for r, row_data in enumerate(legend, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = Font(name='Arial', size=10, bold=(r==1))
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 80

out_path = '/home/z/my-project/download/Мышцы_пробный_фрагмент_v2.xlsx'
wb.save(out_path)
print(f'Saved: {out_path}')
print(f'Rows: {num-1}')
